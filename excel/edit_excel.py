# -*- coding: utf-8 -*-
# ! /usr/local/bin/python3

__version__ = '1.0'
__author__ = 'p.olifer'

from openpyxl import load_workbook
import os
import app_config as config
import printers.printers as p
from datetime import datetime
import shutil
import time


class EditExcelTemplate:
    def __init__(self, template_name):
        now = datetime.now()
        report_name = "_{}{}{}_{}{}{}".format(now.year, now.month, now.day,
                                              now.hour, now.minute, now.second)
        self.report_path = config.EXCEL_REPORT_PATH.format(template_name +
                                                           report_name)

        shutil.copy(src=config.EXCEL_TEMPLATE_PATH.format(template_name),
                    dst=self.report_path)

        # self.wb = load_workbook(filename=config.EXCEL_TEMPLATE_PATH.format(template_name))
        start_load = time.time()
        self.wb = load_workbook(filename=self.report_path,
                                keep_links=False,
                                keep_vba=False,
                                data_only=True)
        end_load = time.time()
        print('LOAD WORKBOOK|{}'.format(str(end_load - start_load)))
        self.ws = self.wb.active
        self.answer = {'file_name': template_name.upper()}

    def write_workbook(self, row_dest, column_dest, value):
        c = self.ws.cell(row=row_dest, column=column_dest)
        c.value = value

    def write_workbook_style(self, row_dest, column_dest, value, style):
        c = self.ws.cell(row=row_dest, column=column_dest)
        c.value = value
        c._style = style

    def add_row(self, row):
        self.ws.insert_rows(row)

    def insert_image(self, image_path, cell):
        self.ws.add_image(img=image_path, anchor=cell)

    def save_excel(self) :
        self.wb.save(self.report_path)

    def print_excel(self, printer_no):
        p.print_excel(printer_no=printer_no, path_to_file=self.report_path)

    def print_excel_file(self, printer_name):
        p.print_excel_file(printer_name=printer_name, path_to_file=self.report_path)

    def delete_file(self, path_to_file):
        p.delete_file(path_to_file=path_to_file)

    def set_answer(self, key, value):
        self.answer[key] = value

    def get_answer(self):
        return self.answer

if __name__ == '__main__':
    pass